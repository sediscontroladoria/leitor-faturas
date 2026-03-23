import io
import os
import sys
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image
from services.processador_dados import ProcessadorDados

class ExportadorRelatorios:
    @staticmethod
    def resource_path(relative_path):
        if hasattr(sys, '_MEIPASS'):
            return os.path.join(sys._MEIPASS, relative_path)
        return os.path.join(os.path.abspath('.'), relative_path)

    @staticmethod
    def gerar_csv(df: pd.DataFrame) -> bytes:
        return df.to_csv(index=False, sep=';').encode('utf-8-sig')

    @staticmethod
    def gerar_excel_relatorio(df_relatorio: pd.DataFrame, faturas_lidas: list, tipo_fatura: str, mes_comp: str, ano_comp: str, tipo_debito: str, conta_fatura: str, complemento: str) -> bytes:
        vl_total = sum(ProcessadorDados._converter_para_float(f.valor) for f in faturas_lidas)
        ir_total = sum(ProcessadorDados._converter_para_float(f.retencao_ir) for f in faturas_lidas)
        vb_total = vl_total + ir_total
        venc_ref = faturas_lidas[0].vencimento if faturas_lidas else 'N/A'
        
        titulo = f'Relatório das faturas {tipo_fatura.upper()} referentes a {mes_comp}/{ano_comp}'
        subtitulo = f'{tipo_debito} - {conta_fatura} - Vencimento {venc_ref} - {complemento}'

        wb = Workbook()
        ws = wb.active
        ws.title = 'Relatório'

        fill_title = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
        fill_header = PatternFill(start_color='DAE3F3', end_color='DAE3F3', fill_type='solid')
        font_title = Font(name='Arial', size=20, color='FFFFFF', bold=True)
        font_header = Font(name='Arial', size=11, bold=True)
        font_normal = Font(name='Arial', size=11)
        align_center = Alignment(horizontal='center', vertical='center')
        border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        ws.row_dimensions[1].height = 75
        ws.merge_cells('A1:H1')
        
        fill_white = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        for col_idx in range(1, 9):
            ws.cell(row=1, column=col_idx).fill = fill_white

        caminho_logo = ExportadorRelatorios.resource_path(os.path.join('assets', 'logo.png'))
        caminho_texto = ExportadorRelatorios.resource_path(os.path.join('assets', 'texto_prefeitura.png'))

        if os.path.exists(caminho_logo):
            img_logo = Image(caminho_logo)
            img_logo.height = 80 
            img_logo.width = 80
            ws.add_image(img_logo, 'A1')

        if os.path.exists(caminho_texto):
            img_texto = Image(caminho_texto)
            img_texto.height = 60 
            img_texto.width = 450
            ws.add_image(img_texto, 'C1')

        ws.merge_cells('A2:H2')
        ws.merge_cells('A3:H3')
        ws['A2'] = titulo
        ws['A3'] = subtitulo
        ws.row_dimensions[2].height = 30
        ws.row_dimensions[3].height = 30

        for row_idx in [2, 3]:
            for col_idx in range(1, 9):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = fill_title
                if col_idx == 1:
                    cell.font = font_title
                    cell.alignment = align_center

        headers_excel = ['DOTAÇÃO', 'AÇÃO', 'SECRETARIA RESPONSÁVEL', 'EMPENHO', 'AF', 'VALOR LÍQUIDO', 'IR', 'VALOR BRUTO']
        ws.append(headers_excel)
        for col_idx in range(1, 9):
            cell = ws.cell(row=4, column=col_idx)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center

        for _, row in df_relatorio.iterrows():
            ws.append(row.tolist())
            current_row = ws.max_row
            for col_idx in range(1, 9):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.font = font_normal
                cell.alignment = align_center
                if col_idx in [6, 7, 8]:
                    cell.number_format = 'R$ #,##0.00'

        last_row = ws.max_row + 1
        ws.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=5)
        ws.cell(row=last_row, column=1, value='Total Geral')
        
        totals = [vl_total, ir_total, vb_total]
        for col_idx in range(1, 9):
            cell = ws.cell(row=last_row, column=col_idx)
            cell.fill = fill_header
            if col_idx <= 5:
                cell.font = font_header
                cell.alignment = align_center
            else:
                cell.value = totals[col_idx - 6]
                cell.font = font_header
                cell.number_format = 'R$ #,##0.00'

        widths = [12, 10, 45, 15, 12, 18, 15, 18]
        col_letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
        for letter, w in zip(col_letters, widths):
            ws.column_dimensions[letter].width = w
            
        for row in ws.iter_rows(min_row=2, max_row=last_row, min_col=1, max_col=8):
            for cell in row:
                cell.border = border_thin

        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        return excel_buffer.getvalue()