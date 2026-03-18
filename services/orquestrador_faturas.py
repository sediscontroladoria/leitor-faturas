import os
import io
import tempfile
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image
from services.factory import ServiceFactory
from services.processador_dados import ProcessadorDados
from services.organizador_faturas import OrganizadorFaturas

class OrquestradorFaturas:
    @staticmethod
    def processar_lote(
        arquivos_pdf, tipo_fatura, mes_comp, ano_comp, 
        tipo_debito, conta_fatura, complemento, 
        mapa_fichas, headers, coluna_id, 
        opcoes_saida, tracker=None
    ):
        leitor = ServiceFactory.get_leitor(tipo_fatura)
        separador = ServiceFactory.get_separador(tipo_fatura)
        organizador = OrganizadorFaturas()

        if not leitor or not separador:
            raise ValueError(f"Ferramentas para {tipo_fatura} não implementadas.")

        with tempfile.TemporaryDirectory() as pasta_trabalho:
            pasta_uploads = os.path.join(pasta_trabalho, 'uploads')
            pasta_separadas = os.path.join(pasta_trabalho, 'separadas')
            pasta_organizadas = os.path.join(pasta_trabalho, 'organizadas')
            os.makedirs(pasta_uploads)

            caminhos_pdfs = []
            for pdf in arquivos_pdf:
                caminho = os.path.join(pasta_uploads, pdf.name)
                with open(caminho, 'wb') as f:
                    f.write(pdf.getbuffer())
                caminhos_pdfs.append(caminho)

            faturas_lidas = []
            for i, caminho_pdf in enumerate(caminhos_pdfs):
                if tracker:
                    tracker.update(i, len(caminhos_pdfs), os.path.basename(caminho_pdf))
                
                fatura = leitor.processar(caminho_pdf)
                
                if tipo_debito == 'Geral':
                    faturas_lidas.append(fatura)
                elif tipo_debito == 'Débito Automático' and fatura.debito_automatico == 'SIM':
                    faturas_lidas.append(fatura)
                elif tipo_debito == 'Débito Manual' and fatura.debito_automatico == 'NÃO':
                    faturas_lidas.append(fatura)

            if not faturas_lidas:
                return None

            resultados = {}

            if 'Gerar ZIP' in opcoes_saida:
                separador.separar(caminhos_pdfs, pasta_separadas)
                organizador.agrupar_por_mapeamento(
                    faturas_lidas, pasta_separadas, pasta_organizadas, mapa_fichas
                )
                zip_path_base = os.path.join(pasta_trabalho, f'Faturas_{tipo_fatura}')
                caminho_zip = organizador.compactar_saida(pasta_organizadas, zip_path_base)
                with open(caminho_zip, 'rb') as f:
                    resultados['zip'] = f.read()

            if 'Gerar Planilha' in opcoes_saida:
                df_final = ProcessadorDados.preparar_dataframe_faturas(
                    faturas_lidas, mapa_fichas, headers, coluna_id
                )
                csv_buffer = io.StringIO()
                df_final.to_csv(csv_buffer, index=False, sep=';', encoding='utf-8-sig')
                resultados['csv'] = csv_buffer.getvalue().encode('utf-8-sig')

            if 'Gerar Relatório Final' in opcoes_saida:
                df_relatorio = ProcessadorDados.gerar_relatorio_final(
                    faturas_lidas, mapa_fichas, coluna_id
                )
                
                vl_total = sum(ProcessadorDados._converter_para_float(f.valor) for f in faturas_lidas)
                ir_total = sum(ProcessadorDados._converter_para_float(f.retencao_ir) for f in faturas_lidas)
                vb_total = vl_total + ir_total
                venc_ref = faturas_lidas[0].vencimento if faturas_lidas else 'N/A'
                
                titulo = f'Relatório das faturas {tipo_fatura.upper()} referentes a {mes_comp}/{ano_comp}'
                subtitulo = f'{tipo_debito} - {conta_fatura} - Vencimento {venc_ref} - {complemento}'

                wb = Workbook()
                ws = wb.active
                ws.title = "Relatório"

                fill_title = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
                fill_header = PatternFill(start_color="DAE3F3", end_color="DAE3F3", fill_type="solid")
                font_title = Font(name="Arial", size=20, color="FFFFFF", bold=True)
                font_header = Font(name="Arial", size=11, bold=True)
                font_normal = Font(name="Arial", size=11)
                align_center = Alignment(horizontal="center", vertical="center")
                border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

                ws.row_dimensions[1].height = 75

                ws.merge_cells('A1:H1')
                
                fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                for col_idx in range(1, 9):
                    ws.cell(row=1, column=col_idx).fill = fill_white

                caminho_logo = os.path.join('assets', 'logo.png')
                caminho_texto = os.path.join('assets', 'texto_prefeitura.png')

                if os.path.exists(caminho_logo):
                    img_logo = Image(caminho_logo)
                    img_logo.height = 80 
                    img_logo.width = 80
                    ws.add_image(img_logo, 'A1')

                if os.path.exists(caminho_texto):
                    img_texto = Image(caminho_texto)
                    img_texto.height = 60 
                    img_texto.width = 450
                    ws.add_image(img_texto, 'C1')

                ws.merge_cells('A2:H2')
                ws.merge_cells('A3:H3')
                ws['A2'] = titulo
                ws['A3'] = subtitulo
                ws.row_dimensions[2].height = 30
                ws.row_dimensions[3].height = 30

                for row_idx in [2, 3]:
                    for col_idx in range(1, 9):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.fill = fill_title
                        if col_idx == 1:
                            cell.font = font_title
                            cell.alignment = align_center

                headers_excel = ['DOTAÇÃO', 'AÇÃO', 'SECRETARIA RESPONSÁVEL', 'EMPENHO', 'AF', 'VALOR LÍQUIDO', 'IR', 'VALOR BRUTO']
                ws.append(headers_excel)
                for col_idx in range(1, 9):
                    cell = ws.cell(row=4, column=col_idx)
                    cell.font = font_header
                    cell.fill = fill_header
                    cell.alignment = align_center

                for _, row in df_relatorio.iterrows():
                    ws.append(row.tolist())
                    current_row = ws.max_row
                    for col_idx in range(1, 9):
                        cell = ws.cell(row=current_row, column=col_idx)
                        cell.font = font_normal
                        cell.alignment = align_center
                        if col_idx in [6, 7, 8]:
                            cell.number_format = 'R$ #,##0.00'

                last_row = ws.max_row + 1
                ws.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=5)
                ws.cell(row=last_row, column=1, value='Total Geral')
                
                totals = [vl_total, ir_total, vb_total]
                for col_idx in range(1, 9):
                    cell = ws.cell(row=last_row, column=col_idx)
                    cell.fill = fill_header
                    if col_idx <= 5:
                        cell.font = font_header
                        cell.alignment = align_center
                    else:
                        cell.value = totals[col_idx - 6]
                        cell.font = font_header
                        cell.number_format = 'R$ #,##0.00'

                widths = [12, 10, 45, 15, 12, 18, 15, 18]
                col_letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
                for letter, w in zip(col_letters, widths):
                    ws.column_dimensions[letter].width = w
                    
                for row in ws.iter_rows(min_row=2, max_row=last_row, min_col=1, max_col=8):
                    for cell in row:
                        cell.border = border_thin

                excel_buffer = io.BytesIO()
                wb.save(excel_buffer)
                resultados['relatorio'] = excel_buffer.getvalue()

            return resultados